from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


DATASET_ID = "faculty_securing_rate"
SCHEMA_VERSION = "v1"
RAW_INPUT_DIR = Path("data/raw") / DATASET_ID / "original"

DETAIL_OUTPUT = (
    f"data/processed/{DATASET_ID}/"
    f"faculty_securing_rate_2015_2025_{SCHEMA_VERSION}_utf8.csv"
)
TOTAL_OUTPUT = (
    f"data/processed/{DATASET_ID}/"
    f"faculty_securing_rate_total_2015_2025_{SCHEMA_VERSION}_utf8.csv"
)
LONG_OUTPUT = (
    f"data/processed/{DATASET_ID}/"
    f"faculty_securing_metric_values_2015_2025_{SCHEMA_VERSION}_utf8.csv"
)
SOURCE_OUTPUT = f"data/metadata/{DATASET_ID}_{SCHEMA_VERSION}.source.json"
SCHEMA_OUTPUT = f"data/metadata/{DATASET_ID}_{SCHEMA_VERSION}_schema.md"
REPORT_OUTPUT = f"data/metadata/{DATASET_ID}_{SCHEMA_VERSION}.processing_report.json"
RAW_README_OUTPUT = f"data/raw/{DATASET_ID}/README.md"

BASE_COLUMNS = [
    "reference_year",
    "survey_round",
    "school_code",
    "campus_type",
    "university_name",
    "field_category",
    "school_type",
    "region_name",
    "founding_type",
    "source_file_name",
]

METRIC_COLUMNS = [
    "교원확보율(전임교원)(편제정원)",
    "교원확보율(전임교원)(재학생)",
    "교원확보율(겸임포함)(편제정원)",
    "교원확보율(겸임포함)(재학생)",
    "교원확보율(초빙포함)(편제정원)",
    "교원확보율(초빙포함)(재학생)",
    "겸임교원확보율(편제정원)",
    "겸임교원확보율(재학생)",
    "겸임교원확보율(편제정원_최종)",
    "겸임교원확보율(재학생_최종)",
]

OUTPUT_COLUMNS = BASE_COLUMNS + METRIC_COLUMNS

SOURCE_METRIC_MAP = {
    "교원확보율(전임교원)(편제정원)": ("full_time_faculty_quota_basis", "교원확보율(전임교원)(편제정원)", "source", ""),
    "교원확보율(전임교원)(재학생)": ("full_time_faculty_enrolled_basis", "교원확보율(전임교원)(재학생)", "source", ""),
    "교원확보율(겸임포함)(편제정원)": ("faculty_with_adjunct_quota_basis", "교원확보율(겸임포함)(편제정원)", "source", ""),
    "교원확보율(겸임포함)(재학생)": ("faculty_with_adjunct_enrolled_basis", "교원확보율(겸임포함)(재학생)", "source", ""),
    "교원확보율(초빙포함)(편제정원)": ("faculty_with_invited_quota_basis", "교원확보율(초빙포함)(편제정원)", "source", ""),
    "교원확보율(초빙포함)(재학생)": ("faculty_with_invited_enrolled_basis", "교원확보율(초빙포함)(재학생)", "source", ""),
    "겸임교원확보율(편제정원)": (
        "adjunct_faculty_quota_basis",
        "겸임교원확보율(편제정원)",
        "derived",
        "교원확보율(겸임포함)(편제정원) - 교원확보율(전임교원)(편제정원)",
    ),
    "겸임교원확보율(재학생)": (
        "adjunct_faculty_enrolled_basis",
        "겸임교원확보율(재학생)",
        "derived",
        "교원확보율(겸임포함)(재학생) - 교원확보율(전임교원)(재학생)",
    ),
    "겸임교원확보율(편제정원_최종)": (
        "adjunct_faculty_quota_basis_final",
        "겸임교원확보율(편제정원_최종)",
        "derived_capped",
        "min(겸임교원확보율(편제정원) * 0.3, 4.0)",
    ),
    "겸임교원확보율(재학생_최종)": (
        "adjunct_faculty_enrolled_basis_final",
        "겸임교원확보율(재학생_최종)",
        "derived_capped",
        "min(겸임교원확보율(재학생) * 0.3, 4.0)",
    ),
}

REQUIRED_HEADER_TARGETS = {
    "교원확보율(전임교원)(편제정원)": ("교원확보율", "전임교원", "편제"),
    "교원확보율(전임교원)(재학생)": ("교원확보율", "전임교원", "재학생"),
    "교원확보율(겸임포함)(편제정원)": ("교원확보율", "겸임포함", "편제"),
    "교원확보율(겸임포함)(재학생)": ("교원확보율", "겸임포함", "재학생"),
    "교원확보율(초빙포함)(편제정원)": ("교원확보율", "초빙", "편제"),
    "교원확보율(초빙포함)(재학생)": ("교원확보율", "초빙", "재학생"),
}

BASE_HEADER_TARGETS = {
    "survey_round": ("조사회차",),
    "school_code": ("학교코드",),
    "campus_type": ("본분교",),
    "university_name": ("학교명",),
    "field_category": ("계열구분", "구분"),
    "school_type": ("학제",),
    "region_name": ("시도",),
    "founding_type": ("설립구분",),
}

COLUMN_DESCRIPTIONS = {
    "reference_year": "기준연도. 조사회차 또는 파일명에서 추출한 4자리 연도",
    "survey_round": "원자료 조사회차",
    "school_code": "원자료 학교코드. 2018년 파일은 해당 컬럼이 없어 결측",
    "campus_type": "원자료 본분교 구분",
    "university_name": "원자료 학교명. 2015년처럼 학교명에 본교가 포함된 값은 원문 보존",
    "field_category": "계열구분/구분. 총계 행이 업로드용 기본 행",
    "school_type": "학제",
    "region_name": "시도",
    "founding_type": "설립구분",
    "source_file_name": "원본 XLSX 파일명",
    "교원확보율(전임교원)(편제정원)": "원자료 교원확보율(%) > 전임교원 > 편제/편제 정원",
    "교원확보율(전임교원)(재학생)": "원자료 교원확보율(%) > 전임교원 > 재학생",
    "교원확보율(겸임포함)(편제정원)": "원자료 교원확보율(%) > 겸임포함 > 편제/편제 정원",
    "교원확보율(겸임포함)(재학생)": "원자료 교원확보율(%) > 겸임포함 > 재학생",
    "교원확보율(초빙포함)(편제정원)": "원자료 교원확보율(%) > 초빙포함 또는 초빙교원 > 편제/편제 정원",
    "교원확보율(초빙포함)(재학생)": "원자료 교원확보율(%) > 초빙포함 또는 초빙교원 > 재학생",
    "겸임교원확보율(편제정원)": "파생값: 겸임포함 편제정원 기준 확보율 - 전임교원 편제정원 기준 확보율",
    "겸임교원확보율(재학생)": "파생값: 겸임포함 재학생 기준 확보율 - 전임교원 재학생 기준 확보율",
    "겸임교원확보율(편제정원_최종)": "파생값: min(겸임교원확보율(편제정원) * 0.3, 4.0)",
    "겸임교원확보율(재학생_최종)": "파생값: min(겸임교원확보율(재학생) * 0.3, 4.0)",
}


def normalize_name(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def compact(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def display_value(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_year_from_name(file_name: str) -> int:
    match = re.search(r"(20\d{2})", file_name)
    if not match:
        raise ValueError(f"파일명에서 연도를 찾을 수 없습니다: {file_name}")
    return int(match.group(1))


def parse_evaluation_cycle(reference_year: int) -> int:
    if reference_year <= 2017:
        return 2
    if reference_year <= 2024:
        return 3
    return 4


def merged_cell_lookup(sheet: Any) -> dict[tuple[int, int], Any]:
    lookup: dict[tuple[int, int], Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = value
    return lookup


def cell_value(sheet: Any, merged_lookup: dict[tuple[int, int], Any], row: int, col: int) -> Any:
    value = sheet.cell(row, col).value
    if value is not None:
        return value
    return merged_lookup.get((row, col))


def header_parts(sheet: Any, merged_lookup: dict[tuple[int, int], Any], col: int) -> tuple[str, str, str]:
    return tuple(display_value(cell_value(sheet, merged_lookup, row, col)) for row in (1, 2, 3))


def compact_header(sheet: Any, merged_lookup: dict[tuple[int, int], Any], col: int) -> str:
    return ">".join(part for part in (compact(cell_value(sheet, merged_lookup, row, col)) for row in (1, 2, 3)) if part)


def human_header(sheet: Any, merged_lookup: dict[tuple[int, int], Any], col: int) -> str:
    return " > ".join(part for part in header_parts(sheet, merged_lookup, col) if part)


def find_base_column(sheet: Any, merged_lookup: dict[tuple[int, int], Any], labels: tuple[str, ...]) -> int | None:
    label_set = {compact(label) for label in labels}
    for col in range(1, sheet.max_column + 1):
        parts = {compact(cell_value(sheet, merged_lookup, row, col)) for row in (1, 2, 3)}
        if parts & label_set:
            return col
    return None


def find_metric_column(
    sheet: Any,
    merged_lookup: dict[tuple[int, int], Any],
    terms: tuple[str, ...],
) -> tuple[int, str]:
    compact_terms = tuple(compact(term) for term in terms)
    matches: list[tuple[int, str]] = []
    for col in range(1, sheet.max_column + 1):
        header = compact_header(sheet, merged_lookup, col)
        if all(term in header for term in compact_terms):
            matches.append((col, human_header(sheet, merged_lookup, col)))
    if len(matches) != 1:
        raise ValueError(
            f"요청 지표 컬럼 매핑 실패: terms={terms}, matches={matches}"
        )
    return matches[0]


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalized_text(value: Any) -> str:
    if value is None:
        return ""
    return display_value(value)


def count_blank_trailing_columns(sheet: Any, merged_lookup: dict[tuple[int, int], Any]) -> int:
    blank_count = 0
    for col in range(sheet.max_column, 0, -1):
        header = compact_header(sheet, merged_lookup, col)
        data_has_value = any(sheet.cell(row, col).value not in (None, "") for row in range(4, sheet.max_row + 1))
        if not header and not data_has_value:
            blank_count += 1
            continue
        break
    return blank_count


def load_source_files(project_root: Path) -> tuple[list[Path], list[str]]:
    raw_dir = project_root / RAW_INPUT_DIR
    if not raw_dir.exists():
        raise FileNotFoundError(f"원자료 디렉토리를 찾을 수 없습니다: {raw_dir}")

    included: list[Path] = []
    excluded: list[str] = []
    for path in raw_dir.glob("*.xlsx"):
        name = normalize_name(path.name)
        if name.startswith("~$"):
            excluded.append(name)
            continue
        if "교원확보율" in name and re.search(r"20\d{2}", name):
            included.append(path)

    # Uploaded Excel lock files can remain in the project root even though only
    # actual source workbooks are copied into data/raw. Record them explicitly
    # so the processing report explains why they were ignored.
    for path in project_root.glob("*.xlsx"):
        name = normalize_name(path.name)
        if name.startswith("~$") and "교원확보율" in name:
            excluded.append(name)

    included.sort(key=lambda item: parse_year_from_name(normalize_name(item.name)))
    excluded = sorted(set(excluded))
    return included, excluded


def process_workbook(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    file_name = normalize_name(path.name)
    reference_year = parse_year_from_name(file_name)
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    merged_lookup = merged_cell_lookup(sheet)

    base_columns = {
        key: find_base_column(sheet, merged_lookup, labels)
        for key, labels in BASE_HEADER_TARGETS.items()
    }
    if base_columns["university_name"] is None or base_columns["field_category"] is None:
        raise ValueError(f"필수 기본 컬럼을 찾지 못했습니다: {file_name}, {base_columns}")

    metric_columns: dict[str, tuple[int, str]] = {
        output_column: find_metric_column(sheet, merged_lookup, terms)
        for output_column, terms in REQUIRED_HEADER_TARGETS.items()
    }

    records: list[dict[str, Any]] = []
    blank_rows_dropped = 0
    for row in range(4, sheet.max_row + 1):
        university_name = normalized_text(cell_value(sheet, merged_lookup, row, base_columns["university_name"]))
        field_category = normalized_text(cell_value(sheet, merged_lookup, row, base_columns["field_category"]))
        if not university_name or not field_category:
            blank_rows_dropped += 1
            continue

        record: dict[str, Any] = {
            "reference_year": reference_year,
            "survey_round": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["survey_round"])) if base_columns["survey_round"] else "",
            "school_code": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["school_code"])) if base_columns["school_code"] else "",
            "campus_type": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["campus_type"])) if base_columns["campus_type"] else "",
            "university_name": university_name,
            "field_category": field_category,
            "school_type": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["school_type"])) if base_columns["school_type"] else "",
            "region_name": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["region_name"])) if base_columns["region_name"] else "",
            "founding_type": normalized_text(cell_value(sheet, merged_lookup, row, base_columns["founding_type"])) if base_columns["founding_type"] else "",
            "source_file_name": file_name,
        }
        for output_column, (column_index, _) in metric_columns.items():
            record[output_column] = to_number(cell_value(sheet, merged_lookup, row, column_index))
        records.append(record)

    frame = pd.DataFrame(records)
    if frame.empty:
        raise ValueError(f"데이터 행이 없습니다: {file_name}")

    frame["겸임교원확보율(편제정원)"] = (
        frame["교원확보율(겸임포함)(편제정원)"]
        - frame["교원확보율(전임교원)(편제정원)"]
    ).round(10)
    frame["겸임교원확보율(재학생)"] = (
        frame["교원확보율(겸임포함)(재학생)"]
        - frame["교원확보율(전임교원)(재학생)"]
    ).round(10)
    frame["겸임교원확보율(편제정원_최종)"] = frame["겸임교원확보율(편제정원)"].apply(
        lambda value: round(min(float(value) * 0.3, 4.0), 10) if pd.notna(value) else pd.NA
    )
    frame["겸임교원확보율(재학생_최종)"] = frame["겸임교원확보율(재학생)"].apply(
        lambda value: round(min(float(value) * 0.3, 4.0), 10) if pd.notna(value) else pd.NA
    )

    mapping_report = {
        "source_file_name": file_name,
        "sheet_name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "data_rows_loaded": int(len(frame)),
        "blank_rows_dropped": int(blank_rows_dropped),
        "blank_trailing_columns": int(count_blank_trailing_columns(sheet, merged_lookup)),
        "missing_base_columns": [key for key, col in base_columns.items() if col is None],
        "base_columns": {
            key: {
                "column_index": col,
                "header": human_header(sheet, merged_lookup, col) if col else None,
            }
            for key, col in base_columns.items()
        },
        "metric_columns": {
            output_column: {
                "column_index": column_index,
                "header": header,
            }
            for output_column, (column_index, header) in metric_columns.items()
        },
    }
    return frame[OUTPUT_COLUMNS], mapping_report


def build_long_frame(detail: pd.DataFrame) -> pd.DataFrame:
    long_records: list[dict[str, Any]] = []
    for source_column in METRIC_COLUMNS:
        metric_id, metric_label, value_source, formula = SOURCE_METRIC_MAP[source_column]
        subset_columns = BASE_COLUMNS + [source_column]
        for row_dict in detail[subset_columns].to_dict(orient="records"):
            value = row_dict[source_column]
            if pd.isna(value):
                continue
            record = {column: row_dict[column] for column in BASE_COLUMNS}
            record.update(
                {
                    "metric_id": metric_id,
                    "metric_label_ko": metric_label,
                    "value": value,
                    "value_source": value_source,
                    "calculation_formula": formula,
                    "unit": "%",
                }
            )
            long_records.append(record)
    long_columns = [
        "metric_id",
        "metric_label_ko",
        "reference_year",
        "survey_round",
        "school_code",
        "campus_type",
        "university_name",
        "field_category",
        "school_type",
        "region_name",
        "founding_type",
        "value",
        "value_source",
        "calculation_formula",
        "unit",
        "source_file_name",
    ]
    return pd.DataFrame(long_records, columns=long_columns)


def sample_negative_rows(frame: pd.DataFrame, column: str) -> list[dict[str, Any]]:
    negative = frame[frame[column] < 0]
    sample_columns = [
        "reference_year",
        "university_name",
        "field_category",
        column,
        "source_file_name",
    ]
    return negative[sample_columns].head(20).to_dict(orient="records")


def validate_calculations(frame: pd.DataFrame) -> dict[str, Any]:
    adjunct_quota_expected = (
        frame["교원확보율(겸임포함)(편제정원)"]
        - frame["교원확보율(전임교원)(편제정원)"]
    ).round(10)
    adjunct_enrolled_expected = (
        frame["교원확보율(겸임포함)(재학생)"]
        - frame["교원확보율(전임교원)(재학생)"]
    ).round(10)
    quota_final_expected = adjunct_quota_expected.apply(lambda value: round(min(float(value) * 0.3, 4.0), 10) if pd.notna(value) else pd.NA)
    enrolled_final_expected = adjunct_enrolled_expected.apply(lambda value: round(min(float(value) * 0.3, 4.0), 10) if pd.notna(value) else pd.NA)

    checks = []
    for name, actual, expected in (
        ("adjunct_quota_difference", frame["겸임교원확보율(편제정원)"], adjunct_quota_expected),
        ("adjunct_enrolled_difference", frame["겸임교원확보율(재학생)"], adjunct_enrolled_expected),
        ("adjunct_quota_final", frame["겸임교원확보율(편제정원_최종)"], quota_final_expected),
        ("adjunct_enrolled_final", frame["겸임교원확보율(재학생_최종)"], enrolled_final_expected),
    ):
        diff = (pd.to_numeric(actual, errors="coerce") - pd.to_numeric(expected, errors="coerce")).abs()
        compared = int(diff.notna().sum())
        checks.append(
            {
                "check_name": name,
                "compared_rows": compared,
                "mismatch_rows_over_1e_9": int((diff > 1e-9).sum()),
                "max_abs_diff": None if compared == 0 else float(diff.max()),
            }
        )
    return {"formula_checks": checks}


def write_outputs(project_root: Path, detail: pd.DataFrame, mappings: list[dict[str, Any]], source_files: list[Path], excluded_files: list[str]) -> dict[str, Any]:
    total = detail[detail["field_category"] == "총계"].copy().reset_index(drop=True)
    long = build_long_frame(detail)

    for relative_path in (DETAIL_OUTPUT, TOTAL_OUTPUT, LONG_OUTPUT, SOURCE_OUTPUT, SCHEMA_OUTPUT, REPORT_OUTPUT, RAW_README_OUTPUT):
        (project_root / relative_path).parent.mkdir(parents=True, exist_ok=True)

    detail.to_csv(project_root / DETAIL_OUTPUT, index=False, encoding="utf-8-sig")
    total.to_csv(project_root / TOTAL_OUTPUT, index=False, encoding="utf-8-sig")
    long.to_csv(project_root / LONG_OUTPUT, index=False, encoding="utf-8-sig")

    processed_at = date.today().isoformat()
    source_payload = {
        "dataset_id": DATASET_ID,
        "dataset_name_ko": "교원확보율",
        "dataset_name_en": "Faculty Securing Rate",
        "schema_version": SCHEMA_VERSION,
        "source_name": "한국대학평가원 대학통계",
        "source_org": "한국대학교육협의회 병설 한국대학평가원",
        "source_url": "https://aims.kcue.or.kr/kor/sub03/eval/evalView.do",
        "source_section": "대학통계 > 대학 및 평가지표 > 데이터 다운로드",
        "download_method": "사용자 수동 업로드 후 로컬 변환",
        "original_file_format": "xlsx",
        "raw_input_directory": RAW_INPUT_DIR.as_posix(),
        "collected_years": sorted(detail["reference_year"].dropna().astype(int).unique().tolist()),
        "downloaded_at": "2026-04-27",
        "processed_at": processed_at,
        "original_files": [normalize_name(path.name) for path in source_files],
        "excluded_files": [
            {"file_name": file_name, "reason": "Excel temporary lock file; excluded from source processing"}
            for file_name in excluded_files
        ],
        "processed_output_files": [DETAIL_OUTPUT, TOTAL_OUTPUT, LONG_OUTPUT],
        "processing_summary": [
            "1~3행 병합 다단 헤더를 의미 기반으로 해제",
            "계열구분/구분, 초빙교원/초빙포함, 편제/편제 정원 표기 차이 정규화",
            "요청한 6개 원천 교원확보율 컬럼을 연도별 위치와 무관하게 추출",
            "겸임교원확보율 파생 컬럼 2개와 0.3 가중 및 4.0 cap 최종 컬럼 2개 생성",
            "전체 계열 상세 CSV, 총계 업로드용 CSV, long metric CSV를 UTF-8-SIG로 저장",
        ],
        "recommended_upload_file": TOTAL_OUTPUT,
        "recommended_detail_file": DETAIL_OUTPUT,
        "recommended_metric_values_file": LONG_OUTPUT,
        "notes": [
            "겸임교원확보율은 원자료 직접 항목이 아니라 겸임포함 확보율에서 전임교원 확보율을 뺀 파생값이다.",
            "최종값은 파생 겸임교원확보율에 0.3을 곱한 값과 4.0 중 작은 값이다.",
            "학교명 표준화는 적용하지 않았고 원자료 표기를 보존했다.",
        ],
    }
    (project_root / SOURCE_OUTPUT).write_text(json.dumps(source_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    negative_columns = ["겸임교원확보율(편제정원)", "겸임교원확보율(재학생)"]
    report = {
        "dataset_id": DATASET_ID,
        "schema_version": SCHEMA_VERSION,
        "processed_at": processed_at,
        "raw_file_count": len(source_files),
        "excluded_temp_file_count": len(excluded_files),
        "excluded_files": [
            {"file_name": file_name, "reason": "Excel temporary lock file; excluded from source processing"}
            for file_name in excluded_files
        ],
        "detail_output": DETAIL_OUTPUT,
        "total_output": TOTAL_OUTPUT,
        "long_output": LONG_OUTPUT,
        "detail_rows": int(len(detail)),
        "total_rows": int(len(total)),
        "detail_columns": int(len(detail.columns)),
        "long_rows": int(len(long)),
        "year_detail_counts": {str(key): int(value) for key, value in detail["reference_year"].value_counts().sort_index().items()},
        "year_total_counts": {str(key): int(value) for key, value in total["reference_year"].value_counts().sort_index().items()},
        "field_category_counts": {str(key): int(value) for key, value in detail["field_category"].value_counts().sort_index().items()},
        "column_mappings": mappings,
        "missing_by_column": {column: int(detail[column].isna().sum()) for column in OUTPUT_COLUMNS},
        "negative_derived_values": {
            column: {
                "count": int((detail[column] < 0).sum()),
                "samples": sample_negative_rows(detail, column),
            }
            for column in negative_columns
        },
        "final_cap_counts": {
            "겸임교원확보율(편제정원_최종)_equals_4": int((detail["겸임교원확보율(편제정원_최종)"] == 4.0).sum()),
            "겸임교원확보율(재학생_최종)_equals_4": int((detail["겸임교원확보율(재학생_최종)"] == 4.0).sum()),
        },
        "validation": validate_calculations(detail),
        "structure_notes": [
            "2015년 파일은 학제/시도/설립구분 컬럼이 없고 학교명에 본교 표기가 포함된 값이 존재한다.",
            "2016년 파일은 데이터 하단에 빈 행 2개와 계열명만 있는 보조 행 7개가 있어 로딩 시 제외한다.",
            "2018년 파일은 학교코드 컬럼이 없어 school_code를 결측으로 보존한다.",
            "2023년 파일은 후행 빈 컬럼 2개가 있어 의미 컬럼 매핑에서 제외한다.",
        ],
    }
    (project_root / REPORT_OUTPUT).write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    schema_lines = [
        "# 교원확보율 v1 스키마",
        "",
        "## 산출물",
        "",
        f"- 상세 CSV: `{DETAIL_OUTPUT}`",
        f"- 업로드용 총계 CSV: `{TOTAL_OUTPUT}`",
        f"- Long metric CSV: `{LONG_OUTPUT}`",
        f"- Source metadata: `{SOURCE_OUTPUT}`",
        f"- Processing report: `{REPORT_OUTPUT}`",
        "",
        "## 컬럼 정의",
        "",
        "| column | description |",
        "| --- | --- |",
    ]
    for column in OUTPUT_COLUMNS:
        schema_lines.append(f"| `{column}` | {COLUMN_DESCRIPTIONS[column]} |")
    schema_lines.extend(
        [
            "",
            "## 처리 로직",
            "",
            "- 원자료 1~3행 병합 헤더를 해제하고, 헤더 텍스트를 공백 제거 후 의미 기반으로 매핑한다.",
            "- `계열구분`과 `구분`, `초빙교원`과 `초빙포함`, `편제`와 `편제 정원`은 동일 의미로 정규화한다.",
            "- `겸임교원확보율(편제정원)` = `교원확보율(겸임포함)(편제정원)` - `교원확보율(전임교원)(편제정원)`.",
            "- `겸임교원확보율(재학생)` = `교원확보율(겸임포함)(재학생)` - `교원확보율(전임교원)(재학생)`.",
            "- `겸임교원확보율(편제정원_최종)` = min(`겸임교원확보율(편제정원)` * 0.3, 4.0).",
            "- `겸임교원확보율(재학생_최종)` = min(`겸임교원확보율(재학생)` * 0.3, 4.0).",
            "- 음수 차이값은 임의로 0 보정하지 않고 그대로 저장하며, 처리 보고서에 건수와 샘플을 남긴다.",
            "",
            "## 사용 권장",
            "",
            "- 시스템 업로드 기본 파일은 `field_category == \"총계\"`만 포함한 총계 CSV를 사용한다.",
            "- 계열별 검증과 추적에는 상세 CSV를 사용한다.",
            "- 학교명 표준화, 통폐합, 캠퍼스 매핑은 이 변환 단계에서 추정하지 않는다.",
        ]
    )
    (project_root / SCHEMA_OUTPUT).write_text("\n".join(schema_lines) + "\n", encoding="utf-8")

    readme = f"""# 교원확보율 원자료

## 위치

- 원자료: `{RAW_INPUT_DIR.as_posix()}`
- 변환 스크립트: `scripts/build_faculty_securing_rate.py`

## 원자료 관리 규칙

- `2015년교원확보율.xlsx`부터 `2025년교원확보율.xlsx`까지 실제 원자료 11개만 처리한다.
- `~$`로 시작하는 Excel 임시 잠금 파일은 처리하지 않는다.
- 원자료 파일명은 연도 추출에 사용하므로 4자리 연도를 유지한다.

## 재생성 명령

```bash
python scripts/build_faculty_securing_rate.py
```

## 산출물

- 상세 CSV: `{DETAIL_OUTPUT}`
- 업로드용 총계 CSV: `{TOTAL_OUTPUT}`
- Long metric CSV: `{LONG_OUTPUT}`
- 스키마: `{SCHEMA_OUTPUT}`
- 출처: `{SOURCE_OUTPUT}`
- 처리 보고서: `{REPORT_OUTPUT}`
"""
    (project_root / RAW_README_OUTPUT).write_text(readme, encoding="utf-8")
    return report


def build(project_root: Path) -> dict[str, Any]:
    source_files, excluded_files = load_source_files(project_root)
    if len(source_files) != 11:
        raise ValueError(f"교원확보율 원자료는 11개여야 합니다. found={len(source_files)}")

    frames: list[pd.DataFrame] = []
    mappings: list[dict[str, Any]] = []
    for path in source_files:
        frame, mapping = process_workbook(path)
        frames.append(frame)
        mappings.append(mapping)

    detail = pd.concat(frames, ignore_index=True)
    detail = detail.sort_values(["reference_year", "university_name", "field_category"]).reset_index(drop=True)
    return write_outputs(project_root, detail, mappings, source_files, excluded_files)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build faculty securing rate CSV package from yearly XLSX files.")
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Project root containing data/raw/faculty_securing_rate/original.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    report = build(args.project_root.resolve())
    print(
        json.dumps(
            {
                "detail_rows": report["detail_rows"],
                "total_rows": report["total_rows"],
                "long_rows": report["long_rows"],
                "outputs": [DETAIL_OUTPUT, TOTAL_OUTPUT, LONG_OUTPUT],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
